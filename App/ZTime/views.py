from django.shortcuts import render, HttpResponse
from django.http import JsonResponse
from App.ZTime.forms import *
from App.ZTime.conexion import *
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
import json
import pandas as pd
import mimetypes
import os
from django.http import HttpResponse, Http404
from wsgiref.util import FileWrapper
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.drawing.image import Image

##LOGIN
from django.contrib.auth.decorators import login_required

def fechaNombre(fecha):
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    now = datetime.now()
    #stringFecha = str(now.day) + "-" + str(now.month) + "-" + str(now.year)
    di = datetime.strptime(str(fecha), "%d/%m/%Y")
    dianum = now.day
    mes = meses[now.month]
    año = now.year
    diaNombre = dias[di.weekday()]
    #fecha = (diaNombre + ", " + str(dianum) + " de " + str(mes) + " del " + str(año))
    return diaNombre

### RENDERIZADO 
@login_required
def zetoneTime(request):
    return render (request, 'ZTime/inicio/index.html')

@login_required
def renderCalcHoras(resquest):
    return render (resquest, 'ZTime/registros/viewRegister.html')

@login_required
def renderVerRegistros(resquest):
    return render (resquest, 'ZTime/registros/ver.html')

@login_required
def renderProcesarRegistros(resquest):
    return render (resquest, 'ZTime/configuracion/procesos.html')


def pruebaHTML(request):
    return render (request, 'prueba/prueba.html')

def prueba(request):
    try:
        ZT = ZetoneTime()
        cursorZT = ZT.cursor()
        sql4 =("SELECT        numero_bulto\n" +
                "FROM            servidordb.Trazabilidad.dbo.Bulto\n" +
                "WHERE Id_bulto = (SELECT MAX(Id_bulto) FROM servidordb.Trazabilidad.dbo.Bulto)")
        cursorZT.execute(sql4)
        consultaZT = cursorZT.fetchall()
        if consultaZT:
            for i in consultaZT:
                print(i[0])
                data = {'bultos': str(i[0])}
                print(data)
            
            cursorZT.commit()
            cursorZT.close()
            ZT.close()
        return JsonResponse(data)
    except Exception as e:
        print("Error")
        print(e)
        data = {'bultos': 'error'}
        return JsonResponse(data)

@csrf_exempt
def calculoHorasJson(request):
    form = form_ver_registros(request.POST)
    if form.is_valid():
        legajo = form.cleaned_data['legajo']
        departamento = form.cleaned_data['departamento']
        desde = form.cleaned_data['desde']
        hasta = form.cleaned_data['hasta']
        desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
        hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
        try:
            registro = []
            ZT = ZetoneTime()
            cursorZT = ZT.cursor()
            sql4 = ("SELECT Legajo, Nombre, Fecha, F1, F2, F3, F4, HorasF1F2, HorasF3F4, HorasDeMas\n" +
                    "FROM TemporalHoras\n" +
                    "WHERE Legajo = '" + str(legajo) + "' AND FechaHora >= '"+ str(desdeSql) +"' AND FechaHora <= '"+ str(hastaSql) +"'\n"+
                    "ORDER BY Legajo, FechaHora")
            cursorZT.execute(sql4)
            consultaZT = cursorZT.fetchall()
            if consultaZT:
                for i in consultaZT:
                    dia = fechaNombre(str(i[2]))
                    resultado = {'legajo': str(i[0]), 'nombre': i[1], 'dia': dia, 'fecha': i[2], 'f1': i[3] or "-", 'f2': i[4] or "-", 'f3': i[5] or "-", 'f4': i[6] or "-", 'hm':i[7] or "-", 'ht': i[8] or "-", 'ex': i[9] or "-"}
                    registro.append(resultado)

                jsonList = json.dumps({'message': 'Success', 'registros':registro}) 
                return JsonResponse(jsonList, safe=False)
            else:
                jsonList = json.dumps({'message':'No se encontraron fichadas.'}) 
                return JsonResponse(jsonList, safe=False)
        except Exception as e:
            print(e)
            error = 'Error: ' + str(e)
            jsonList = json.dumps({'message':error}) 
            return JsonResponse(jsonList, safe=False)
        finally:
            cursorZT.close()
            ZT.close()
    else:
        departamento = form.cleaned_data['departamento']
        desde = form.cleaned_data['desde']
        hasta = form.cleaned_data['hasta']
        desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
        hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
        if departamento == "Todos":
            try:
                registro = []
                ZT = ZetoneTime()
                cursorZT = ZT.cursor()
                sql4 = ("SELECT Legajo, Nombre, Fecha, F1, F2, F3, F4, HorasF1F2, HorasF3F4, HorasDeMas\n" +
                        "FROM TemporalHoras\n" +
                        "WHERE FechaHora >= '"+ str(desdeSql) +"' AND FechaHora <= '"+ str(hastaSql) +"'\n"+
                        "ORDER BY Legajo, FechaHora")
                cursorZT.execute(sql4)
                consultaZT = cursorZT.fetchall()
                if consultaZT:
                    for i in consultaZT:
                        dia = fechaNombre(str(i[2]))
                        resultado = {'legajo': str(i[0]), 'nombre': i[1], 'dia': dia, 'fecha': i[2], 'f1': i[3] or "-", 'f2': i[4] or "-", 'f3': i[5] or "-", 'f4': i[6] or "-", 'hm':i[7] or "-", 'ht': i[8] or "-", 'ex': i[9] or "-"}
                        registro.append(resultado)

                    jsonList = json.dumps({'message': 'Success', 'registros':registro}) 
                    return JsonResponse(jsonList, safe=False)
                else:
                    jsonList = json.dumps({'message':'No se encontraron fichadas'}) 
                    return JsonResponse(jsonList, safe=False)
            except Exception as e:
                print(e)
                error = 'Error: ' + str(e)
                jsonList = json.dumps({'message':error}) 
                return JsonResponse(jsonList, safe=False)
            finally:
                cursorZT.close()
                ZT.close()
        else:
            jsonList = json.dumps({'message':'Debe seleccionar un Departamento'}) 
            return JsonResponse(jsonList, safe=False)

@csrf_exempt
def ver_registros_sin_proceso(request):
    form = form_ver_registros(request.POST)
    if form.is_valid():
        departamento = form.cleaned_data['departamento']
        legajo = form.cleaned_data['legajo']
        desde = form.cleaned_data['desde']
        hasta = form.cleaned_data['hasta']
        desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
        hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
        try:
            registro = []
            ZT = ZetoneTime()
            cursorZT = ZT.cursor()
            sql_consulta = ("SELECT DISTINCT\n" +
                                         "LEFT(CONVERT(varchar, iclock_transaction_1.punch_time, 108), 2) AS soloHora, iclock_transaction_1.emp_code AS Legajo, LEFT(Legajos.Nombre, 28) AS Nombre, CONVERT(varchar(10), iclock_transaction_1.punch_time, 103) AS Fecha, \n" +
                                         "CONVERT(varchar, iclock_transaction_1.punch_time, 108) AS Hora, iclock_transaction_1.punch_time AS FechaHora\n" +
                         "FROM            servidordb.zkbiotime.dbo.iclock_transaction AS iclock_transaction_1 INNER JOIN\n" +
                                                 "Legajos ON iclock_transaction_1.emp_code = Legajos.Legajos\n" +
                         "WHERE      iclock_transaction_1.emp_code = '"+str(legajo)+"' AND TRY_CONVERT(DATE, iclock_transaction_1.punch_time) >= '"+str(desdeSql)+"' AND TRY_CONVERT(DATE, iclock_transaction_1.punch_time) <= '"+str(hastaSql)+"' \n" +
                         "ORDER BY Legajo, iclock_transaction_1.punch_time")
            cursorZT.execute(sql_consulta)
            consulta = cursorZT.fetchall()
            if consulta:
                for i in consulta:
                    dia = fechaNombre(str(i[3]))
                    resultado = {'legajo': str(i[1]), 'nombre': str(i[2]), 'dia': dia, 'fecha':str(i[3]), 'hora': str(i[4])}
                    registro.append(resultado)
                #print(registro)
                jsonList = json.dumps({'message':'Success', 'registros': registro}) 
                return JsonResponse(jsonList, safe=False)
            else:
                try:
                    registro = []
                    ZT2 = ZetoneTime()
                    cursorZT2 = ZT2.cursor()
                    sql_consulta = ("SELECT        hr_employee_1.emp_pin AS Legajo, hr_employee_1.emp_lastname + ' ' + hr_employee_1.emp_firstname AS Nombre,CONVERT(varchar(10), att_punches_1.punch_time, 103) AS Fecha, CONVERT(varchar, att_punches_1.punch_time, 108) AS Hora \n" +
                                    "FROM            servidordb.ZKTime.dbo.att_punches AS att_punches_1 INNER JOIN\n" +
                                                            "servidordb.ZKTime.dbo.hr_employee AS hr_employee_1 ON att_punches_1.employee_id = hr_employee_1.id\n" +
                                    "WHERE hr_employee_1.emp_pin = '"+ str(legajo) +"' AND TRY_CONVERT(DATE, att_punches_1.punch_time) >= '"+ str(desdeSql) +"' AND TRY_CONVERT(DATE, att_punches_1.punch_time) <= '"+str(hastaSql)+"'\n" +
                                    "ORDER BY hr_employee_1.emp_lastname, hr_employee_1.emp_pin, att_punches_1.punch_time")
                    cursorZT2.execute(sql_consulta)
                    consulta = cursorZT2.fetchall()
                    if consulta:
                        for i in consulta:
                            dia = fechaNombre(str(i[2]))
                            resultado = {'legajo': str(i[0]), 'nombre': str(i[1]), 'dia': dia, 'fecha':str(i[2]), 'hora': str(i[3])}
                            registro.append(resultado)
                        #print(registro)
                        jsonList = json.dumps({'message':'Success', 'registros': registro}) 
                        return JsonResponse(jsonList, safe=False)
                    else:
                        jsonList = json.dumps({'message':'No se encontraron fichadas'}) 
                        return JsonResponse(jsonList, safe=False)
                except Exception as e:
                    print(e)
                    error = 'Error: ' + str(e)
                    jsonList = json.dumps({'message':error}) 
                    return JsonResponse(jsonList, safe=False)
                finally:
                    cursorZT2.close()
                    ZT2.close()
        except Exception as e:
            error = 'Error: ' + str(e)
            jsonList = json.dumps({'message':error}) 
            return JsonResponse(jsonList, safe=False)
        finally:
            cursorZT.close()
            ZT.close()
    else:
        departamento = form.cleaned_data['departamento']
        if departamento == "Todos":
            #legajo = form.cleaned_data['legajo',0]
            desde = form.cleaned_data['desde']
            hasta = form.cleaned_data['hasta']
            desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
            hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
            try:
                registro = []
                ZT = ZetoneTime()
                cursorZT = ZT.cursor()
                sql_consulta = ("SELECT DISTINCT\n" +
                                            "LEFT(CONVERT(varchar, iclock_transaction_1.punch_time, 108), 2) AS soloHora, iclock_transaction_1.emp_code AS Legajo, LEFT(Legajos.Nombre, 28) AS Nombre, CONVERT(varchar(10), iclock_transaction_1.punch_time, 103) AS Fecha, \n" +
                                            "CONVERT(varchar, iclock_transaction_1.punch_time, 108) AS Hora, iclock_transaction_1.punch_time AS FechaHora\n" +
                            "FROM            servidordb.zkbiotime.dbo.iclock_transaction AS iclock_transaction_1 INNER JOIN\n" +
                                                    "Legajos ON iclock_transaction_1.emp_code = Legajos.Legajos\n" +
                            "WHERE      TRY_CONVERT(DATE, iclock_transaction_1.punch_time) >= '"+str(desdeSql)+"' AND TRY_CONVERT(DATE, iclock_transaction_1.punch_time) <= '"+str(hastaSql)+"' \n" +
                            "ORDER BY Legajo, iclock_transaction_1.punch_time")
                cursorZT.execute(sql_consulta)
                consulta = cursorZT.fetchall()
                if consulta:
                    for i in consulta:
                        dia = fechaNombre(str(i[3]))
                        resultado = {'legajo': str(i[1]), 'nombre': str(i[2]), 'dia': dia, 'fecha':str(i[3]), 'hora': str(i[4])}
                        registro.append(resultado)
                    #print(registro)
                    jsonList = json.dumps({'message':'Success', 'registros': registro}) 
                    return JsonResponse(jsonList, safe=False)
                else:
                    jsonList = json.dumps({'message':'No se encontraron fichadas'}) 
                    return JsonResponse(jsonList, safe=False)
            except Exception as e:
                print(e)
                error = 'Error: ' + str(e)
                jsonList = json.dumps({'message':error}) 
                return JsonResponse(jsonList, safe=False)
            finally:
                cursorZT.close()
                ZT.close()
        elif departamento == "Administración":
            desde = form.cleaned_data['desde']
            hasta = form.cleaned_data['hasta']
            desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
            hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
            try:
                registro = []
                ZT = ZetoneTime()
                cursorZT = ZT.cursor()
                sql_consulta = ("SELECT        hr_employee_1.emp_pin AS Legajo, hr_employee_1.emp_lastname + ' ' + hr_employee_1.emp_firstname AS Nombre,CONVERT(varchar(10), att_punches_1.punch_time, 103) AS Fecha, CONVERT(varchar, att_punches_1.punch_time, 108) AS Hora \n" +
                                "FROM            servidordb.ZKTime.dbo.att_punches AS att_punches_1 INNER JOIN\n" +
                                                        "servidordb.ZKTime.dbo.hr_employee AS hr_employee_1 ON att_punches_1.employee_id = hr_employee_1.id\n" +
                                "WHERE TRY_CONVERT(DATE, att_punches_1.punch_time) >= '"+ str(desdeSql) +"' AND TRY_CONVERT(DATE, att_punches_1.punch_time) <= '"+str(hastaSql)+"'\n" +
                                "ORDER BY hr_employee_1.emp_lastname, hr_employee_1.emp_pin, att_punches_1.punch_time")
                cursorZT.execute(sql_consulta)
                consulta = cursorZT.fetchall()
                if consulta:
                    for i in consulta:
                        dia = fechaNombre(str(i[2]))
                        resultado = {'legajo': str(i[0]), 'nombre': str(i[1]), 'dia': dia, 'fecha':str(i[2]), 'hora': str(i[3])}
                        registro.append(resultado)
                    #print(registro)
                    jsonList = json.dumps({'message':'Success', 'registros': registro}) 
                    return JsonResponse(jsonList, safe=False)
                else:
                    jsonList = json.dumps({'message':'No se encontraron fichadas'}) 
                    return JsonResponse(jsonList, safe=False)
            except Exception as e:
                print(e)
                error = 'Error: ' + str(e)
                jsonList = json.dumps({'message':error}) 
                return JsonResponse(jsonList, safe=False)
            finally:
                cursorZT.close()
                ZT.close()
        else:
            jsonList = json.dumps({'message':'Debe seleccionar un Departamento'}) 
            return JsonResponse(jsonList, safe=False)

@csrf_exempt
def excelCreateRegistros(request):
    form = form_ver_registros(request.POST)
    if form.is_valid():
        departamento = form.cleaned_data['departamento']
        legajo = form.cleaned_data['legajo']
        desde = form.cleaned_data['desde']
        hasta = form.cleaned_data['hasta']
        desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
        hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
        try:
            ZT = ZetoneTime()
            cursorZT = ZT.cursor()
            sql_consulta = ("SELECT DISTINCT\n" +
                                         "LEFT(CONVERT(varchar, iclock_transaction_1.punch_time, 108), 2) AS soloHora, iclock_transaction_1.emp_code AS Legajo, LEFT(Legajos.Nombre, 28) AS Nombre, CONVERT(varchar(10), iclock_transaction_1.punch_time, 103) AS Fecha, \n" +
                                         "CONVERT(varchar, iclock_transaction_1.punch_time, 108) AS Hora, iclock_transaction_1.punch_time AS FechaHora\n" +
                         "FROM            servidordb.zkbiotime.dbo.iclock_transaction AS iclock_transaction_1 INNER JOIN\n" +
                                                 "Legajos ON iclock_transaction_1.emp_code = Legajos.Legajos\n" +
                         "WHERE      iclock_transaction_1.emp_code = '"+str(legajo)+"' AND TRY_CONVERT(DATE, iclock_transaction_1.punch_time) >= '"+str(desdeSql)+"' AND TRY_CONVERT(DATE, iclock_transaction_1.punch_time) <= '"+str(hastaSql)+"' \n" +
                         "ORDER BY Legajo, iclock_transaction_1.punch_time")
            cursorZT.execute(sql_consulta)
            consulta = cursorZT.fetchall()
            if consulta:
                legajos = []
                nombres = []
                dias = []
                fechas = []
                horas = []
                fechaHoras = []
                for i in consulta:
                    dia = fechaNombre(str(i[3]))
                    legajos.append(i[1])
                    nombres.append(i[2])
                    dias.append(dia)
                    fechas.append(i[3])
                    horas.append(i[4])
                    fechaHoras.append(i[5])
                cantidad = int (len(legajos))
                book = Workbook()
                sheet = book.active
                borde = Side(border_style='thin', color='000000')
                bordes = Border(top=borde, left=borde, bottom=borde, right=borde)
                logo = Image('App/ZTime/data/logos/Zetone.png')
                sheet.add_image(logo, 'B1')
                sheet['B4'] = "Legajo"
                sheet['B4'].font = Font(bold=True)
                sheet['B4'].border = bordes
                sheet['C4'] = "Nombre y Apellido"
                sheet['C4'].font = Font(bold=True)
                sheet['C4'].border = bordes
                sheet['D4'] = "Día"
                sheet['D4'].font = Font(bold=True)
                sheet['D4'].border = bordes
                sheet['E4'] = "Fecha"
                sheet['E4'].font = Font(bold=True)
                sheet['E4'].border = bordes
                sheet['F4'] = "Hora"
                sheet['F4'].font = Font(bold=True)
                sheet['F4'].border = bordes
                sheet['G4'] = "Fecha y Hora"
                sheet['G4'].font = Font(bold=True)
                sheet['G4'].border = bordes
                numero = 0
                for j in range (5,(cantidad + 5)):
                    sheet[f'B{j}'] = legajos[numero]
                    sheet[f'B{j}'].border = bordes
                    sheet[f'C{j}'] = nombres[numero]
                    sheet[f'C{j}'].border = bordes
                    sheet[f'D{j}'] = dias[numero]
                    sheet[f'D{j}'].border = bordes
                    sheet[f'E{j}'] = fechas[numero]
                    sheet[f'E{j}'].border = bordes
                    sheet[f'F{j}'] = str(horas[numero])
                    sheet[f'F{j}'].border = bordes
                    sheet[f'G{j}'] = fechaHoras[numero]
                    sheet[f'G{j}'].border = bordes
                    numero = numero + 1
                now = datetime.now().time()
                nowHour = str(now.hour) + "_" + str(now.minute) + "_" + str(now.second)
                nombre_excel ="Reporte_Registros_Legajo_" + str(legajos[0]) + "_" + str(desde) + "_" + str(hasta) + "_" + nowHour + ".xlsx"
                book.save('App/ZTime/data/excel/' + nombre_excel)
                jsonList = json.dumps({'message':'Success', 'excel': nombre_excel}) 
                return JsonResponse(jsonList, safe=False)
            else:
                jsonList = json.dumps({'message':'No se encontraron fichadas'}) 
                return JsonResponse(jsonList, safe=False)
        except Exception as e:
            print(e)
            error = 'Error: ' + str(e)
            jsonList = json.dumps({'message':error}) 
            return JsonResponse(jsonList, safe=False)
        finally:
            cursorZT.close()
            ZT.close()
    else:
        departamento = form.cleaned_data['departamento']
        if departamento == "Todos":
            #legajo = form.cleaned_data['legajo',0]
            desde = form.cleaned_data['desde']
            hasta = form.cleaned_data['hasta']
            desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
            hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
            try:
                dbZetoneTime = ZetoneTime()
                cursorZTime = dbZetoneTime.cursor()
                consultaSql = ("SELECT DISTINCT \n" +
                                                "LEFT(CONVERT(varchar, iclock_transaction_1.punch_time, 108), 2) AS soloHora, iclock_transaction_1.emp_code AS Legajo, LEFT(Legajos.Nombre, 28) AS Nombre, CONVERT(varchar(10), iclock_transaction_1.punch_time, 103) AS Fecha,\n" +
                                                "CONVERT(varchar, iclock_transaction_1.punch_time, 108) AS Hora, iclock_transaction_1.punch_time AS FechaHora\n" +
                        "FROM            servidordb.zkbiotime.dbo.iclock_transaction AS iclock_transaction_1 INNER JOIN\n" +
                                                "Legajos ON iclock_transaction_1.emp_code = Legajos.Legajos\n" +
                        "WHERE      TRY_CONVERT(DATE, iclock_transaction_1.punch_time) >= '"+ desdeSql +"' AND TRY_CONVERT(DATE, iclock_transaction_1.punch_time) <= '"+ hastaSql +"' \n" +
                        "ORDER BY Legajo, iclock_transaction_1.punch_time")
                cursorZTime.execute(consultaSql)
                consulta = cursorZTime.fetchall()
                if consulta:
                    legajos = []
                    nombres = []
                    dias = []
                    fechas = []
                    horas = []
                    fechaHoras = []
                    for i in consulta:
                        dia = fechaNombre(str(i[3]))
                        legajos.append(i[1])
                        nombres.append(i[2])
                        dias.append(dia)
                        fechas.append(i[3])
                        horas.append(i[4])
                        fechaHoras.append(i[5])
                    
                ##CREO LAS FUNCIONES  CREAR UN EXCEL 
                    cantidad = int (len(legajos))
                    book = Workbook()
                    sheet = book.active
                    borde = Side(border_style='thin', color='000000')
                    bordes = Border(top=borde, left=borde, bottom=borde, right=borde)
                    logo = Image('App/ZTime/data/logos/Zetone.png')
                    sheet.add_image(logo, 'B1')
                    sheet['B4'] = "Legajo"
                    sheet['B4'].font = Font(bold=True)
                    sheet['B4'].border = bordes
                    sheet['C4'] = "Nombre y Apellido"
                    sheet['C4'].font = Font(bold=True)
                    sheet['C4'].border = bordes
                    sheet['D4'] = "Día"
                    sheet['D4'].font = Font(bold=True)
                    sheet['D4'].border = bordes
                    sheet['E4'] = "Fecha"
                    sheet['E4'].font = Font(bold=True)
                    sheet['E4'].border = bordes
                    sheet['F4'] = "Hora"
                    sheet['F4'].font = Font(bold=True)
                    sheet['F4'].border = bordes
                    sheet['G4'] = "Fecha y Hora"
                    sheet['G4'].font = Font(bold=True)
                    sheet['G4'].border = bordes
                    numero = 0
                    for j in range (5,(cantidad + 5)):
                        #print (numero)
                        sheet[f'B{j}'] = legajos[numero]
                        sheet[f'B{j}'].border = bordes
                        sheet[f'C{j}'] = nombres[numero]
                        sheet[f'C{j}'].border = bordes
                        sheet[f'D{j}'] = dias[numero]
                        sheet[f'D{j}'].border = bordes
                        sheet[f'E{j}'] = fechas[numero]
                        sheet[f'E{j}'].border = bordes
                        sheet[f'F{j}'] = str(horas[numero])
                        sheet[f'F{j}'].border = bordes
                        sheet[f'G{j}'] = fechaHoras[numero]
                        sheet[f'G{j}'].border = bordes
                        numero = numero + 1
                    now = datetime.now().time()
                    nowHour = str(now.hour) + "_" + str(now.minute) + "_" + str(now.second)
                    nombre_excel ="Reporte_Registros_" + str(desde) + "_" + str(hasta) + "_" + nowHour + ".xlsx"
                    book.save('App/ZTime/data/excel/' + nombre_excel)
                    jsonList = json.dumps({'message':'Success', 'excel': nombre_excel}) 
                    return JsonResponse(jsonList, safe=False)
                else:
                    jsonList = json.dumps({'message':'No se encontraron fichadas'}) 
                    return JsonResponse(jsonList, safe=False)
            except Exception as e:
                print(e)
                error = 'Error: ' + str(e)
                jsonList = json.dumps({'message':error}) 
                return JsonResponse(jsonList, safe=False)
            finally:
                cursorZTime.close()
                dbZetoneTime.close()
        elif departamento == "Administración":
            #legajo = form.cleaned_data['legajo',0]
            desde = form.cleaned_data['desde']
            hasta = form.cleaned_data['hasta']
            desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
            hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
            try:
                dbZetoneTime = ZetoneTime()
                cursorZTime = dbZetoneTime.cursor()
                consultaSql = ("SELECT        hr_employee_1.emp_pin AS Legajo, hr_employee_1.emp_lastname + ' ' + hr_employee_1.emp_firstname AS Nombre,CONVERT(varchar(10), att_punches_1.punch_time, 103) AS Fecha, CONVERT(varchar, att_punches_1.punch_time, 108) AS Hora, att_punches_1.punch_time \n" +
                                "FROM            servidordb.ZKTime.dbo.att_punches AS att_punches_1 INNER JOIN\n" +
                                                        "servidordb.ZKTime.dbo.hr_employee AS hr_employee_1 ON att_punches_1.employee_id = hr_employee_1.id\n" +
                                "WHERE TRY_CONVERT(DATE, att_punches_1.punch_time) >= '"+ str(desdeSql) +"' AND TRY_CONVERT(DATE, att_punches_1.punch_time) <= '"+str(hastaSql)+"'\n" +
                                "ORDER BY hr_employee_1.emp_lastname, hr_employee_1.emp_pin, att_punches_1.punch_time")
                cursorZTime.execute(consultaSql)
                consulta = cursorZTime.fetchall()
                if consulta:
                    legajos = []
                    nombres = []
                    dias = []
                    fechas = []
                    horas = []
                    fechaHoras = []
                    for i in consulta:
                        dia = fechaNombre(str(i[2]))
                        legajos.append(i[0])
                        nombres.append(i[1])
                        dias.append(dia)
                        fechas.append(i[2])
                        horas.append(i[3])
                        fechaHoras.append(i[4])
                    
                ##CREO LAS FUNCIONES  CREAR UN EXCEL 
                    cantidad = int (len(legajos))
                    book = Workbook()
                    sheet = book.active
                    borde = Side(border_style='thin', color='000000')
                    bordes = Border(top=borde, left=borde, bottom=borde, right=borde)
                    logo = Image('App/ZTime/data/logos/Zetone.png')
                    sheet.add_image(logo, 'B1')
                    sheet['B5'] = "Legajo"
                    sheet['B5'].font = Font(bold=True)
                    sheet['B5'].border = bordes
                    sheet['C5'] = "Nombre y Apellido"
                    sheet['C5'].font = Font(bold=True)
                    sheet['C5'].border = bordes
                    sheet['D5'] = "Día"
                    sheet['D5'].font = Font(bold=True)
                    sheet['D5'].border = bordes
                    sheet['E5'] = "Fecha"
                    sheet['E5'].font = Font(bold=True)
                    sheet['E5'].border = bordes
                    sheet['F5'] = "Hora"
                    sheet['F5'].font = Font(bold=True)
                    sheet['F5'].border = bordes
                    sheet['G5'] = "Fecha y Hora"
                    sheet['G5'].font = Font(bold=True)
                    sheet['G5'].border = bordes
                    nowH = datetime.now().time()
                    nowF = datetime.now()
                    fechasExporta = str(nowF.day) + "/" + str(nowF.month) + "/" + str(nowF.year)
                    horaExporta = str(nowH.hour) + ":" + str(nowH.minute) + ":" + str(nowH.second)
                    ##DATOS
                    sheet['F1'] = "Fecha: "
                    sheet['G1'] = fechasExporta
                    sheet['F1'].font = Font(bold=True)
                    sheet['F2'] = "Hora: "
                    sheet['G2'] = horaExporta
                    sheet['F2'].font = Font(bold=True)
                    sheet['F3'] = "Depto o Legajo: "
                    sheet['G3'] = departamento
                    sheet['F3'].font = Font(bold=True)
                    numero = 0
                    for j in range (6,(cantidad + 6)):
                        #print (numero)
                        sheet[f'B{j}'] = legajos[numero]
                        sheet[f'B{j}'].border = bordes
                        sheet[f'C{j}'] = nombres[numero]
                        sheet[f'C{j}'].border = bordes
                        sheet[f'D{j}'] = dias[numero]
                        sheet[f'D{j}'].border = bordes
                        sheet[f'E{j}'] = fechas[numero]
                        sheet[f'E{j}'].border = bordes
                        sheet[f'F{j}'] = str(horas[numero])
                        sheet[f'F{j}'].border = bordes
                        sheet[f'G{j}'] = fechaHoras[numero]
                        sheet[f'G{j}'].border = bordes
                        numero = numero + 1
                    now = datetime.now().time()
                    nowHour = str(now.hour) + "_" + str(now.minute) + "_" + str(now.second)
                    nombre_excel ="Reporte_Registros_" + str(desde) + "_" + str(hasta) + "_" + nowHour + ".xlsx"
                    book.save('App/ZTime/data/excel/' + nombre_excel)
                    jsonList = json.dumps({'message':'Success', 'excel': nombre_excel}) 
                    return JsonResponse(jsonList, safe=False)
                else:
                    jsonList = json.dumps({'message':'No se encontraron fichadas'}) 
                    return JsonResponse(jsonList, safe=False)
            except Exception as e:
                print(e)
                error = 'Error: ' + str(e)
                jsonList = json.dumps({'message':error}) 
                return JsonResponse(jsonList, safe=False)
            finally:
                cursorZTime.close()
                dbZetoneTime.close()
        else:
            jsonList = json.dumps({'message':'Debe Seleccionar un Departamento'}) 
            return JsonResponse(jsonList, safe=False)

##CREACION DE EXCEL CACLCULO DE HORAS
@csrf_exempt
def createExcelCalculo(request):
    form = form_ver_registros(request.POST)
    if form.is_valid():
        legajo = form.cleaned_data['legajo']
        departamento = form.cleaned_data['departamento']
        desde = form.cleaned_data['desde']
        hasta = form.cleaned_data['hasta']
        desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
        hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
        try:
            ZT = ZetoneTime()
            cursorZT = ZT.cursor()
            sql4 = ("SELECT Legajo, Nombre, Fecha, F1, F2, F3, F4, HorasF1F2, HorasF3F4, HorasDeMas\n" +
                    "FROM TemporalHoras\n" +
                    "WHERE Legajo = '" + str(legajo) + "' AND FechaHora >= '"+ str(desdeSql) +"' AND FechaHora <= '"+ str(hastaSql) +"'\n"+
                    "ORDER BY FechaHora")
            cursorZT.execute(sql4)
            consultaZT = cursorZT.fetchall()
            if consultaZT:
                #LISTAS
                legajos = []
                nombres = []
                dias = []
                fechas = []
                f1 = []
                f2 = []
                f3 = []
                f4 = []
                hm = []
                ht = []
                ex = []
                for i in consultaZT:
                    dia = fechaNombre(str(i[2]))
                    legajos.append(i[0])
                    nombres.append(i[1])
                    dias.append(dia)
                    fechas.append(i[2])
                    f1.append(i[3] or "-")
                    f2.append(i[4] or "-")
                    f3.append(i[5] or "-")
                    f4.append(i[6] or "-")
                    hm.append(i[7] or "-")
                    ht.append(i[8] or "-")
                    ex.append(i[9] or "-")

                cantidad = int (len(legajos))
                book = Workbook()
                sheet = book.active
                borde = Side(border_style='thin', color='000000')
                bordes = Border(top=borde, left=borde, bottom=borde, right=borde)
                logo = Image('App/ZTime/data/logos/Zetone.png')
                sheet.add_image(logo, 'B1')
                sheet['B4'] = "Legajo"
                sheet['B4'].font = Font(bold=True)
                sheet['B4'].border = bordes
                sheet['C4'] = "Nombre y Apellido"
                sheet['C4'].font = Font(bold=True)
                sheet['C4'].border = bordes
                sheet['D4'] = "Día"
                sheet['D4'].font = Font(bold=True)
                sheet['D4'].border = bordes
                sheet['E4'] = "Fecha"
                sheet['E4'].font = Font(bold=True)
                sheet['E4'].border = bordes
                sheet['F4'] = "Entrada"
                sheet['F4'].font = Font(bold=True)
                sheet['F4'].border = bordes
                sheet['G4'] = "Salida"
                sheet['G4'].font = Font(bold=True)
                sheet['G4'].border = bordes
                sheet['H4'] = "Entrada"
                sheet['H4'].font = Font(bold=True)
                sheet['H4'].border = bordes
                sheet['I4'] = "Salida"
                sheet['I4'].font = Font(bold=True)
                sheet['I4'].border = bordes
                sheet['J4'] = "Hs. Mañana"
                sheet['J4'].font = Font(bold=True)
                sheet['J4'].border = bordes
                sheet['K4'] = "Hs. Tarde"
                sheet['K4'].font = Font(bold=True)
                sheet['K4'].border = bordes
                sheet['L4'] = "Hs. Extra"
                sheet['L4'].font = Font(bold=True)
                sheet['L4'].border = bordes
                numero = 0
                for j in range (5,(cantidad + 5)):
                    sheet[f'B{j}'] = legajos[numero]
                    sheet[f'B{j}'].border = bordes
                    sheet[f'C{j}'] = nombres[numero]
                    sheet[f'C{j}'].border = bordes
                    sheet[f'D{j}'] = dias[numero]
                    sheet[f'D{j}'].border = bordes
                    sheet[f'E{j}'] = fechas[numero]
                    sheet[f'E{j}'].border = bordes

                    sheet[f'F{j}'] = str(f1[numero])
                    sheet[f'F{j}'].border = bordes
                    sheet[f'G{j}'] = str(f2[numero])
                    sheet[f'G{j}'].border = bordes
                    sheet[f'H{j}'] = str(f3[numero])
                    sheet[f'H{j}'].border = bordes
                    sheet[f'I{j}'] = str(f4[numero])
                    sheet[f'I{j}'].border = bordes
                    sheet[f'J{j}'] = str(hm[numero])
                    sheet[f'J{j}'].border = bordes
                    sheet[f'K{j}'] = str(ht[numero])
                    sheet[f'K{j}'].border = bordes
                    sheet[f'L{j}'] = str(ex[numero])
                    sheet[f'L{j}'].border = bordes
                    numero = numero + 1
                now = datetime.now().time()
                nowHour = str(now.hour) + "_" + str(now.minute) + "_" + str(now.second)
                nombre_excel ="Reporte_Registros_Calculo_Legajo_" + str(legajos[0]) + "_" + str(desde) + "_" + str(hasta) + "_" + nowHour + ".xlsx"
                book.save('App/ZTime/data/excel/' + nombre_excel)
                jsonList = json.dumps({'message':'Success', 'excel': nombre_excel}) 
                return JsonResponse(jsonList, safe=False)
            else:
                jsonList = json.dumps({'message':'No se encontraron fichadas.'}) 
                return JsonResponse(jsonList, safe=False)
        except Exception as e:
            print(e)
            error = 'Error: ' + str(e)
            jsonList = json.dumps({'message':error}) 
            return JsonResponse(jsonList, safe=False)
        finally:
            cursorZT.close()
            ZT.close()
    else:
        departamento = form.cleaned_data['departamento']
        if departamento == "Todos":
            departamento = form.cleaned_data['departamento']
            desde = form.cleaned_data['desde']
            hasta = form.cleaned_data['hasta']
            desdeSql = datetime.strptime(str(desde), "%Y-%m-%d").strftime("%d/%m/%Y")
            hastaSql = datetime.strptime(str(hasta), "%Y-%m-%d").strftime("%d/%m/%Y")
            try:
                ZT = ZetoneTime()
                cursorZT = ZT.cursor()
                sql4 = ("SELECT Legajo, Nombre, Fecha, F1, F2, F3, F4, HorasF1F2, HorasF3F4, HorasDeMas\n" +
                        "FROM TemporalHoras\n" +
                        "WHERE FechaHora >= '"+ str(desdeSql) +"' AND FechaHora <= '"+ str(hastaSql) +"'\n"+
                        "ORDER BY Legajo")
                cursorZT.execute(sql4)
                consultaZT = cursorZT.fetchall()
                if consultaZT:
                    #LISTAS
                    legajos = []
                    nombres = []
                    dias = []
                    fechas = []
                    f1 = []
                    f2 = []
                    f3 = []
                    f4 = []
                    hm = []
                    ht = []
                    ex = []
                    for i in consultaZT:
                        dia = fechaNombre(str(i[2]))
                        legajos.append(i[0])
                        nombres.append(i[1])
                        dias.append(dia)
                        fechas.append(i[2])
                        f1.append(i[3] or "-")
                        f2.append(i[4] or "-")
                        f3.append(i[5] or "-")
                        f4.append(i[6] or "-")
                        hm.append(i[7] or "-")
                        ht.append(i[8] or "-")
                        ex.append(i[9] or "-")

                    cantidad = int (len(legajos))
                    book = Workbook()
                    sheet = book.active
                    borde = Side(border_style='thin', color='000000')
                    bordes = Border(top=borde, left=borde, bottom=borde, right=borde)
                    logo = Image('App/ZTime/data/logos/Zetone.png')
                    sheet.add_image(logo, 'B1')
                    sheet['B4'] = "Legajo"
                    sheet['B4'].font = Font(bold=True)
                    sheet['B4'].border = bordes
                    sheet['C4'] = "Nombre y Apellido"
                    sheet['C4'].font = Font(bold=True)
                    sheet['C4'].border = bordes
                    sheet['D4'] = "Día"
                    sheet['D4'].font = Font(bold=True)
                    sheet['D4'].border = bordes
                    sheet['E4'] = "Fecha"
                    sheet['E4'].font = Font(bold=True)
                    sheet['E4'].border = bordes
                    sheet['F4'] = "Entrada"
                    sheet['F4'].font = Font(bold=True)
                    sheet['F4'].border = bordes
                    sheet['G4'] = "Salida"
                    sheet['G4'].font = Font(bold=True)
                    sheet['G4'].border = bordes
                    sheet['H4'] = "Entrada"
                    sheet['H4'].font = Font(bold=True)
                    sheet['H4'].border = bordes
                    sheet['I4'] = "Salida"
                    sheet['I4'].font = Font(bold=True)
                    sheet['I4'].border = bordes
                    sheet['J4'] = "Hs. Mañana"
                    sheet['J4'].font = Font(bold=True)
                    sheet['J4'].border = bordes
                    sheet['K4'] = "Hs. Tarde"
                    sheet['K4'].font = Font(bold=True)
                    sheet['K4'].border = bordes
                    sheet['L4'] = "Hs. Extra"
                    sheet['L4'].font = Font(bold=True)
                    sheet['L4'].border = bordes
                    numero = 0
                    for j in range (5,(cantidad + 5)):
                        sheet[f'B{j}'] = legajos[numero]
                        sheet[f'B{j}'].border = bordes
                        sheet[f'C{j}'] = nombres[numero]
                        sheet[f'C{j}'].border = bordes
                        sheet[f'D{j}'] = dias[numero]
                        sheet[f'D{j}'].border = bordes
                        sheet[f'E{j}'] = fechas[numero]
                        sheet[f'E{j}'].border = bordes
                        sheet[f'F{j}'] = str(f1[numero])
                        sheet[f'F{j}'].border = bordes
                        sheet[f'G{j}'] = str(f2[numero])
                        sheet[f'G{j}'].border = bordes
                        sheet[f'H{j}'] = str(f3[numero])
                        sheet[f'H{j}'].border = bordes
                        sheet[f'I{j}'] = str(f4[numero])
                        sheet[f'I{j}'].border = bordes
                        sheet[f'J{j}'] = str(hm[numero])
                        sheet[f'J{j}'].border = bordes
                        sheet[f'K{j}'] = str(ht[numero])
                        sheet[f'K{j}'].border = bordes
                        sheet[f'L{j}'] = str(ex[numero])
                        sheet[f'L{j}'].border = bordes
                        numero = numero + 1
                    now = datetime.now().time()
                    nowHour = str(now.hour) + "_" + str(now.minute) + "_" + str(now.second)
                    nombre_excel ="Reporte_Registros_Calculo" + str(desde) + "_" + str(hasta) + "_" + nowHour + ".xlsx"
                    book.save('App/ZTime/data/excel/' + nombre_excel)
                    jsonList = json.dumps({'message':'Success', 'excel': nombre_excel}) 
                    return JsonResponse(jsonList, safe=False)
                else:
                    jsonList = json.dumps({'message':'No se encoontraron fichadas.'}) 
                    return JsonResponse(jsonList, safe=False)
            except Exception as e:
                print(e)
                error = 'Error: ' + str(e)
                jsonList = json.dumps({'message':error}) 
                return JsonResponse(jsonList, safe=False)
            finally:
                cursorZT.close()
                ZT.close()
        else:
            jsonList = json.dumps({'message':'Seleccione que un Departamento.'}) 
            return JsonResponse(jsonList, safe=False)

def download_excel(request, file_path):
    file_path = 'App/ZTime/data/excel/' + file_path
    if os.path.exists(file_path):
        with open(file_path, 'rb') as excel:
            mime_type, _ = mimetypes.guess_type(file_path)
            response = HttpResponse(FileWrapper(excel), content_type=mime_type)
            response['Content-Disposition'] = f"attachment; filename={os.path.basename(file_path)}"
            return response
    raise Http404

def delete_xlsx_files(request):
    try:
        for filename in os.listdir('App/ZTime/data/excel/'):
            if filename.endswith(".xlsx"):
                os.remove(os.path.join('App/ZTime/data/excel/', filename))
        jsonList = json.dumps({'message':'Se Borraron todos los archivos.'}) 
        return JsonResponse(jsonList, safe=False)
    except Exception as e:
        jsonList = json.dumps({'message':'Se produjo un error al borrar los archivos.'}) 
        return JsonResponse(jsonList, safe=False)

def download_Excel(self):
    try:
        f = open('App/ZTime/data/excel/prueba.xlsx', 'rb')
        response = HttpResponse(content=f)
        response['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        print(e)
        respuesta = 'Error'
        lista_estado= [{'Info':e, 'Info2': respuesta}]
        estado = [lista_estado]
        return HttpResponse(estado)


def descargar_archivo(request): 
 
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) 
 
    filename = 'vikosur.pdf'
    filepath = BASE_DIR + '/ZTime/data/excel/' + filename 
 
    path = open(filepath, 'r') 
 
    mime_type, _ = mimetypes.guess_type(filepath)
    
    response = HttpResponse(path, content_type = mime_type)
 
    response['Content-Disposition'] = f"attachment; filename={filename}"
 
    return response

### CAPTURA LEGAJOS
def legajos(fecha):
    try:
        ZT = ZetoneTime()
        cursorZT = ZT.cursor()
        sqlConsulta =("SELECT DISTINCT emp_code FROM servidordb.zkbiotime.dbo.iclock_transaction\n" +
                "WHERE TRY_CONVERT(DATE, punch_time) = '" + fecha + "'")
        cursorZT.execute(sqlConsulta)
        consultaZT = cursorZT.fetchall()
        if consultaZT:
            lista_legajos = []
            for i in consultaZT:
                legajo = str(i[0])
                lista_legajos.append(legajo)
            return lista_legajos
        else:
             return 0
    except Exception as e:
        print(e)
        return 5
    finally:
        cursorZT.close()
        ZT.close()

def grupo_horario(id):
    try:
        ZT = ZetoneTime()
        cursorZT = ZT.cursor()
        sqlConsulta =("SELECT        Grupos_Horarios.Nombre, Horarios.EM1, Horarios.EM2, Horarios.SM1, Horarios.SM2, Horarios.ET1, Horarios.ET2, Horarios.ST1, Horarios.ST2\n" +
                        "FROM            Grupos_Horarios INNER JOIN\n" +
                                                "Horarios ON Grupos_Horarios.ID = Horarios.ID_Grupo\n" +
                        "WHERE        (Grupos_Horarios.ID = '" + str(id) + "')")
        cursorZT.execute(sqlConsulta)
        consultaZT = cursorZT.fetchone()
        if consultaZT:
            lista_horario = []
            index = 0
            for i in consultaZT:
                lista_horario.append(str(i[index]))
                index = index + 1
            return lista_horario
        else:
             return 0
    except Exception as e:
        print(e)
        return 5
    finally:
        cursorZT.close()
        ZT.close()

def tipo_hora(hora):
    horario = datetime.strptime(str(hora), "%X").time()



    entradaMañana1 = datetime.strptime("05:00:00", "%X").time()
    entradaMañana2 = datetime.strptime("09:02:00", "%X").time()

    salidaMañana1 = datetime.strptime("10:55:00", "%X").time()
    salidaMañana2 = datetime.strptime("13:12:00", "%X").time()

    entradaTarde1 = datetime.strptime("13:45:00", "%X").time()
    entradaTarde2 = datetime.strptime("16:55:00", "%X").time()

    salidaTarde1 = datetime.strptime("17:00:00", "%X").time()
    salidaTarde2 = datetime.strptime("21:30:00", "%X").time()

    if entradaMañana1 < horario < entradaMañana2:
        tipo = 'F1'
        return tipo
    if salidaMañana1 < horario < salidaMañana2:
        tipo = 'F2'
        return tipo 
    if entradaTarde1 < horario < entradaTarde2:
        tipo = 'F3'
        return tipo
    if salidaTarde1 < horario < salidaTarde2:
        tipo = 'F4'
        return tipo

def existe_fichada(legajo, fecha):
    try:
        conexion = ZetoneTime()
        cursor_cox = conexion.cursor()
        consulta = ("SELECT * FROM TemporalHoras WHERE Legajo ='" + str(legajo) + "' AND Fecha='" + str(fecha) + "'")
        cursor_cox.execute(consulta)
        existe = cursor_cox.fetchone()
        if existe:
            return 1
        else:
            return 0
    except Exception as e:
        print(e)
        return 5
    finally:
        cursor_cox.close()
        conexion.close()

def horas_fichadas(legajo, fecha):
    try:
        conexion = ZetoneTime()
        cursor_cox = conexion.cursor()
        consultaSQL = ("SELECT CONVERT(varchar, punch_time, 108) AS Hora\n" +
                        "FROM servidordb.zkbiotime.dbo.iclock_transaction\n" +
                        "WHERE TRY_CONVERT(DATE, punch_time) = '" + str(fecha) +"' AND emp_code = '" + str(legajo) + "'")
        cursor_cox.execute(consultaSQL)
        consulta = cursor_cox.fetchall()
        if consulta:
            listado_horas = []
            for i in consulta:
                listado_horas.append(i[0])
            return listado_horas
        else:
            return 0
    except Exception as e:
        print(e)
        return 5
    finally:
        cursor_cox.close()
        conexion.close()

def nombre_legajo(legajo):
    try:
        conexion = ZetoneTime()
        cursor_cox = conexion.cursor()
        consultaSQL = ("SELECT Nombre FROM Legajos WHERE Legajos ='" + str(legajo) + "'")
        cursor_cox.execute(consultaSQL)
        consulta = cursor_cox.fetchone()
        if consulta:
            nombre = consulta[0]
            return nombre
        else:
            nombre = "Sin Asignar"
            return nombre
    except Exception as e:
        print(e)
        return 5
    finally:
        cursor_cox.close()
        conexion.close()

def inserta_fichada(legajo, nombre, fecha, hora, fecha_hora, tipo):
    if tipo != None:
        try:
            update = ZetoneTime()
            cursorUpdate = update.cursor()
            consulta = ("INSERT INTO TemporalHoras (Legajo, Nombre, " + str(tipo) + ", Fecha, FechaHora) VALUES ('" + str(legajo) + "','" + str(nombre) + "','" + str(hora) + "', '" + str(fecha) + "', '" + str(fecha_hora) + "')")
            cursorUpdate.execute(consulta)
            cursorUpdate.commit()
        except Exception as e:
            print(e)
            print("Inserta Fichada " + str(legajo) + " --> " + str(hora))
        finally:
            cursorUpdate.close()
            update.close()

def update_fichada(legajo, fecha, hora, tipo):
    if tipo != None:
        try:
            update = ZetoneTime()
            cursorUpdate = update.cursor()
            sql = ("UPDATE TemporalHoras SET " + str(tipo) + "='" + str(hora) + "' WHERE Legajo='" + str(legajo) +"' AND Fecha='" + str(fecha) + "'")
            #print(sql)
            cursorUpdate.execute(sql)
            cursorUpdate.commit()
        except Exception as e:
            print("Actualizar Hora " + str(legajo) + " --> " + str(hora))
            print(e)
        finally:
            cursorUpdate.close()
            update.close()


## VISTA DE PROCESO DE HORAS POST
def proceso_horas(request):
    grupo = '1'
    form = form_proceso_fichadas(request.POST)
    if form.is_valid():
        f = form.cleaned_data['fecha']
        departamento = form.cleaned_data['departamento']
        fecha = datetime.strptime(str(f), "%Y-%m-%d").strftime("%d/%m/%Y")
        fecha_hora = str(fecha) + " 00:00:00"
        if legajos(str(fecha)) == 0:
            jsonList = json.dumps({'message':'No hay Fichadas para la fecha: ' + str(fecha)}) 
            return JsonResponse(jsonList, safe=False)
        else:
            listado_legajos = legajos(str(fecha))
            for i in listado_legajos:
                legajo = str(i)
                nombre = nombre_legajo(legajo)
                listado_horas = horas_fichadas(legajo,fecha)
                #print(listado_horas)
                for j in listado_horas:
                    hora = str(j)
                    if existe_fichada(legajo,fecha) == 0:
                        ##INSERTA
                        tipo = tipo_hora(hora)
                        inserta_fichada(legajo,nombre,fecha,hora,fecha_hora,tipo)
                    elif existe_fichada(legajo,fecha) == 1:
                        ##ACTUALIZA
                        tipo = tipo_hora(hora)
                        update_fichada(legajo,fecha,hora,tipo)
                    else:
                        jsonList = json.dumps({'message':'Se produjo un error.'}) 
                        return JsonResponse(jsonList, safe=False)
                    
            jsonList = json.dumps({'message':'Success', 'fecha': fecha}) 
            return JsonResponse(jsonList, safe=False)
    else:
        jsonList = json.dumps({'message':'Formulario no válido.'}) 
        return JsonResponse(jsonList, safe=False)


@csrf_exempt
def post_recive_data(request):
    if request.method == 'POST':
        data = request.POST
        data = json.loads(request.body)
        fecha = data['Fecha']
        empresa = data['Empresa']
        accion = data['Accion']
        # aquí puedes procesar los datos recibidos
        print(fecha)
        print(empresa)
        print(accion)
        
        registro = [{"lote": "49255 / LAS ACACIAS", "bins": "38", "hora": "08:12"}, {"lote": "49260 / LAS ACACIAS", "bins": "43", "hora": "08:56"}]
        listado = [{"nombre": "josue", "segundo": "ruben", "apellido": "chambi"}, {"nombre": "gabriela", "segundo": "lidia", "apellido": "astrada"}, {"nombre": "persona x", "segundo": "x", "apellido": "xxx"}]
        return JsonResponse({"message": "ok", "data": listado})
    else:
        return JsonResponse({'message': 'Not found'})
    
